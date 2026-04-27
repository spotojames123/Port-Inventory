"""
build_port_inventory.py — per-switch timestamped XLSXs (dotted MACs + last_output)

Outputs one XLSX PER SWITCH in ./reports named "<switch_name>-YYYYMMDD-HHMMSS.xlsx".

Sheet 1 - Port Inventory:
  switch_name, switch_ip, snmp_location, interface, description, connected,
  trunk_mode, vlans, voice_vlan, duplex, speed, power_watts, max_sec_devices,
  last_input, last_output, end_device_mac, end_device_ip, ip_source

Sheet 2 - Cleanup:
  Summary stats (total ports, used, down, % utilization) +
  List of UP ports idle for more than 4 weeks (cols N & O > 4 weeks)
"""

from __future__ import annotations
import re
import os
import time
import argparse
from typing import Dict, Any, List
from netmiko import ConnectHandler
import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# -------------------------------- Interface name matcher --------------------------------
IF_NAME = re.compile(
    r"^(?:Te|TenGigabitEthernet|Gi|GigabitEthernet|Fa|FastEthernet|Fo|FortyGigabitEthernet|Hu|HundredGigE|Eth|Ethernet|Po|Port-channel|Fi|FiveGigabitEthernet)\d+(?:/\d+){0,3}$",
    re.I,
)

# -------------------------------- Regex Parsers --------------------------------
RE_SNMP_LOC = re.compile(r"Location\s*:\s*(?P<loc>.*)")
RE_SNMP_LOC_RUN = re.compile(r"^snmp-server\s+location\s+(?P<loc>.+)$", re.M)

RE_SWITCHPORT_SECTION_SPLIT = re.compile(r"^Name:\s*", re.M)
RE_SWITCHPORT_MODE = re.compile(r"Administrative Mode:\s*(?P<admin>\S+).*?Operational Mode:\s*(?P<oper>\S+)", re.S)
RE_SWITCHPORT_VLANS = re.compile(
    r"Access Mode VLAN:\s*(?P<access>\S+).*?"
    r"(?:Trunking Native Mode VLAN:\s*(?P<native>\S+))?.*?"
    r"Trunking VLANs Enabled:\s*(?P<trunk>.+?)\n",
    re.S,
)
RE_SWITCHPORT_VOICE = re.compile(r"Voice VLAN:\s*(?P<voice>\S+)", re.S)

RE_INTF_HDR = re.compile(
    r"^(?P<intf>[A-Za-z].+?)\s+is\s+(?P<line_state>administratively down|up|down),\s+line protocol is\s+(?P<proto>up|down)"
)
RE_LAST_INPUT = re.compile(r"Last input\s+(?P<li>[^,]+)")
RE_LAST_IO = re.compile(r"Last input\s+(?P<input>[^,]+),\s+output\s+(?P<output>[^,]+)", re.I)

RE_DESC_ROW = re.compile(r"^(?P<intf>\S+)\s+\S+\s+\S+\s+(?P<desc>.*)$")

FLOAT = re.compile(r"^\d+(?:\.\d+)?$")

RE_MAC_TABLE_RELAXED = re.compile(
    r"^(?P<vlan>\d+)\s+(?P<mac>[0-9a-fA-F\.:\-]+)\s+(?P<type>dynamic|static)\s+(?P<intf>(?:Te|Gi|Fa|Hu|Eth|Po)\S+)$",
    re.I | re.M,
)
RE_ARP_ROW = re.compile(r"Internet\s+(?P<ip>\d+\.\d+\.\d+\.\d+)\s+\d+\s+(?P<mac>[0-9a-fA-F\.:\-]+)")
RE_DHCP_IP_MAC = re.compile(r"^(?P<ip>\d+\.\d+\.\d+\.\d+)\s+(?P<mac>[0-9a-fA-F\.:\-]+)\s+", re.M)
RE_DHCP_MAC_IP = re.compile(r"^(?P<mac>[0-9a-fA-F\.:\-]+)\s+(?P<ip>\d+\.\d+\.\d+\.\d+)\s+", re.M)

RE_SW_STATUS = re.compile(
    r"^(?P<intf>\S+)\s+.*?\s{2,}(?P<status>\S+)\s+(?P<vlan>\S+)\s+(?P<duplex>\S+)\s+(?P<speed>\S+)\s+(?P<type>.+)$",
    re.I,
)

# -------------------------------- Idle time helpers --------------------------------
def parse_idle_seconds(time_str: str) -> float:
    """
    Convert Cisco time strings to seconds.
    Handles: never, 00:00:00, 1d02h, 2w3d, 4w5d, 1y2w, etc.
    Returns float('inf') for 'never', 0 if unparseable.
    """
    if not time_str:
        return 0
    s = time_str.strip().lower()
    if s == "never":
        return float("inf")

    total = 0.0
    # hh:mm:ss
    m = re.match(r"^(\d+):(\d+):(\d+)$", s)
    if m:
        return int(m.group(1)) * 3600 + int(m.group(2)) * 60 + int(m.group(3))

    # Cisco compound: Xy Xw Xd Xh Xm Xs
    for value, unit in re.findall(r"(\d+)([ywdhms])", s):
        v = int(value)
        if unit == "y":
            total += v * 365 * 86400
        elif unit == "w":
            total += v * 7 * 86400
        elif unit == "d":
            total += v * 86400
        elif unit == "h":
            total += v * 3600
        elif unit == "m":
            total += v * 60
        elif unit == "s":
            total += v
    return total

FOUR_WEEKS_SECONDS = 4 * 7 * 86400

def is_idle_over_4_weeks(last_input: str, last_output: str) -> bool:
    """Return True if BOTH last_input and last_output indicate > 4 weeks idle."""
    li = parse_idle_seconds(last_input)
    lo = parse_idle_seconds(last_output)
    return li > FOUR_WEEKS_SECONDS and lo > FOUR_WEEKS_SECONDS

# -------------------------------- MAC helpers --------------------------------
def norm_mac(mac: str) -> str:
    return re.sub(r"[^0-9a-f]", "", mac.lower())

def format_mac(mac: str) -> str:
    flat = norm_mac(mac)
    if len(flat) == 12:
        return f"{flat[0:4]}.{flat[4:8]}.{flat[8:12]}"
    return mac

# -------------------------------- Connections --------------------------------
def connect(dev: Dict[str, Any]):
    return ConnectHandler(
        device_type=dev.get("device_type", "cisco_ios"),
        host=dev["host"],
        username=dev.get("username"),
        password=dev.get("password"),
        secret=dev.get("secret", ""),
        fast_cli=True,
        timeout=25,
    )

# -------------------------------- Collectors & Parsers --------------------------------
def get_snmp_location(conn) -> str:
    try:
        out = conn.send_command("show snmp")
        m = RE_SNMP_LOC.search(out)
        if m:
            return m.group("loc").strip()
    except Exception:
        pass
    try:
        out = conn.send_command("show running-config | include ^snmp-server location ")
        m = RE_SNMP_LOC_RUN.search(out)
        if m:
            return m.group("loc").strip()
    except Exception:
        pass
    return ""

def parse_interfaces_status(text: str) -> Dict[str, Dict[str, Any]]:
    out: Dict[str, Dict[str, Any]] = {}
    seen_header = False
    for raw in text.splitlines():
        line = raw.rstrip("\r")
        if not line.strip():
            continue
        low = line.lower()
        if not seen_header and low.startswith("port") and "status" in low and "duplex" in low:
            seen_header = True
            continue
        if not seen_header:
            continue
        if set(line.strip()) <= set("-=+"):
            continue
        m = RE_SW_STATUS.match(line)
        if not m:
            continue
        intf   = m.group("intf").strip()
        status = m.group("status").strip().lower()
        vlan   = m.group("vlan").strip()
        duplex = m.group("duplex").strip()
        speed  = m.group("speed").strip()
        if not IF_NAME.match(intf):
            continue
        out[intf] = {
            "connected": (status == "connected"),
            "duplex": duplex,
            "speed": speed,
            "vlan_col": vlan,
        }
    return out

def parse_switchport_detail(text: str) -> Dict[str, Dict[str, Any]]:
    out: Dict[str, Dict[str, Any]] = {}
    sections = RE_SWITCHPORT_SECTION_SPLIT.split(text)
    for sec in sections[1:]:
        first_line_end = sec.find("\n")
        intf = sec[:first_line_end].strip()
        if not IF_NAME.match(intf):
            continue
        d = out.setdefault(intf, {})
        m_mode = RE_SWITCHPORT_MODE.search(sec)
        if m_mode:
            d["trunk_mode"] = (m_mode.group("oper").lower() == "trunk")
        m_v = RE_SWITCHPORT_VLANS.search(sec)
        if m_v:
            if d.get("trunk_mode"):
                native = (m_v.group("native") or "").strip()
                trunk = (m_v.group("trunk") or "").strip()
                d["vlans"] = f"native {native}; allowed {trunk}".strip()
            else:
                d["vlans"] = (m_v.group("access") or "").strip()
        m_voice = RE_SWITCHPORT_VOICE.search(sec)
        if m_voice:
            d["voice_vlan"] = (m_voice.group("voice") or "").strip()
    return out

def parse_interfaces_last_input_bulk(text: str) -> Dict[str, Dict[str, Any]]:
    out: Dict[str, Dict[str, Any]] = {}
    current = None
    for line in text.splitlines():
        h = RE_INTF_HDR.match(line)
        if h:
            current = h.group("intf").split()[0]
            if IF_NAME.match(current):
                out.setdefault(current, {})
            else:
                current = None
            continue
        if current is None:
            continue
        m_li = RE_LAST_INPUT.search(line)
        if m_li:
            out[current]["last_input"] = m_li.group("li").strip()
    return out

def get_last_io_per_interface(conn, intf: str) -> tuple[str, str]:
    try:
        txt = conn.send_command(f"show interface {intf} | include Last input")
        m = RE_LAST_IO.search(txt)
        if m:
            return m.group("input").strip(), m.group("output").strip()
    except Exception:
        pass
    return "", ""

def parse_descriptions(text: str) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for line in text.splitlines():
        m = RE_DESC_ROW.match(line.strip())
        if not m:
            continue
        intf = m.group("intf")
        if IF_NAME.match(intf):
            out[intf] = (m.group("desc") or "").strip()
    return out

def parse_poe(text: str) -> Dict[str, float]:
    watts: Dict[str, float] = {}
    for line in text.splitlines():
        parts = line.split()
        if len(parts) < 2:
            continue
        port = parts[0]
        if not IF_NAME.match(port):
            continue
        w = None
        if len(parts) >= 4 and FLOAT.match(parts[3]):
            w = float(parts[3])
        elif FLOAT.match(parts[-1]):
            w = float(parts[-1])
        if w is not None:
            watts[port] = w
    return watts

def parse_mac_table(text: str) -> Dict[str, List[str]]:
    res: Dict[str, List[str]] = {}
    for m in RE_MAC_TABLE_RELAXED.finditer(text or ""):
        intf = m.group("intf")
        mac  = norm_mac(m.group("mac"))
        if IF_NAME.match(intf):
            res.setdefault(intf, []).append(mac)
    for line in (text or "").splitlines():
        ls = line.strip()
        if not ls or ls.lower().startswith(("vlan", "----")):
            continue
        parts = ls.split()
        if len(parts) >= 4 and (parts[2].lower() in ("dynamic", "static")):
            vlan, mac, _typ, port = parts[0], parts[1], parts[2], parts[3]
            if IF_NAME.match(port):
                res.setdefault(port, []).append(norm_mac(mac))
        elif len(parts) >= 3 and IF_NAME.match(parts[-1]) and re.match(r"^\d+$", parts[0]):
            vlan, mac, port = parts[0], parts[1], parts[-1]
            res.setdefault(port, []).append(norm_mac(mac))
    return res

def parse_arp(text: str) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for m in RE_ARP_ROW.finditer(text or ""):
        out[norm_mac(m.group("mac"))] = m.group("ip")
    return out

def parse_dhcp_snoop(text: str) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for m in RE_DHCP_IP_MAC.finditer(text or ""):
        out[norm_mac(m.group("mac"))] = "DHCP"
    for m in RE_DHCP_MAC_IP.finditer(text or ""):
        out[norm_mac(m.group("mac"))] = "DHCP"
    return out

# -------------------------------- XLSX Styles --------------------------------
HEADER_FILL  = PatternFill("solid", start_color="1F4E79")
HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
SUBHDR_FILL  = PatternFill("solid", start_color="2E75B6")
SUBHDR_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
STAT_FILL    = PatternFill("solid", start_color="D6E4F0")
IDLE_FILL    = PatternFill("solid", start_color="FCE4D6")  # orange tint for idle rows
CENTER       = Alignment(horizontal="center", vertical="center")
THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

def style_cell(cell, bold=False, fill=None, font_color="000000", center=False, border=True):
    cell.font = Font(bold=bold, color=font_color, name="Arial", size=10)
    if fill:
        cell.fill = fill
    if center:
        cell.alignment = CENTER
    if border:
        cell.border = THIN_BORDER

# -------------------------------- Write XLSX --------------------------------
def write_xlsx(fname: str, rows: List[Dict[str, Any]], switch_name: str):
    fields = [
        "switch_name","switch_ip","snmp_location","interface","description","connected","trunk_mode","vlans","voice_vlan",
        "duplex","speed","power_watts","max_sec_devices","last_input","last_output","end_device_mac","end_device_ip","ip_source"
    ]

    wb = Workbook()

    # ── Sheet 1: Port Inventory ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Port Inventory"

    # Header row
    for col_idx, field in enumerate(fields, start=1):
        cell = ws1.cell(row=1, column=col_idx, value=field.replace("_", " ").title())
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, field in enumerate(fields, start=1):
            val = row.get(field, "")
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.border = THIN_BORDER
            # Highlight connected column (F = col 6)
            if field == "connected":
                if val == "up":
                    cell.fill = PatternFill("solid", start_color="C6EFCE")
                    cell.font = Font(name="Arial", size=10, color="276221")
                else:
                    cell.fill = PatternFill("solid", start_color="FFCCCC")
                    cell.font = Font(name="Arial", size=10, color="9C0006")

    # Auto-width
    col_widths = {i: len(f.replace("_", " ").title()) for i, f in enumerate(fields, start=1)}
    for row in rows:
        for col_idx, field in enumerate(fields, start=1):
            val = str(row.get(field, ""))
            col_widths[col_idx] = max(col_widths[col_idx], len(val))
    for col_idx, width in col_widths.items():
        ws1.column_dimensions[get_column_letter(col_idx)].width = min(width + 2, 50)

    ws1.freeze_panes = "A2"

    # ── Sheet 2: Cleanup ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Cleanup")

    total_ports = len(rows)
    used_ports  = sum(1 for r in rows if r.get("connected") == "up")
    down_ports  = total_ports - used_ports

    # ── Summary block (row 1–3) ──────────────────────────────────────────────
    summary_headers = ["Total Ports", "Used Ports (Up)", "Down Ports", "Utilization %"]
    for col_idx, hdr in enumerate(summary_headers, start=1):
        hcell = ws2.cell(row=1, column=col_idx, value=hdr)
        hcell.font = HEADER_FONT
        hcell.fill = HEADER_FILL
        hcell.alignment = CENTER
        hcell.border = THIN_BORDER
        ws2.column_dimensions[get_column_letter(col_idx)].width = 20

    # Values
    ws2.cell(row=2, column=1, value=total_ports)
    ws2.cell(row=2, column=2, value=used_ports)
    ws2.cell(row=2, column=3, value=down_ports)
    # Utilization as formula referencing B2/A2 (handles zero-division)
    ws2["D2"] = f'=IFERROR(B2/A2,0)'
    ws2["D2"].number_format = "0.0%"

    for col_idx in range(1, 5):
        cell = ws2.cell(row=2, column=col_idx)
        cell.font = Font(bold=True, name="Arial", size=11)
        cell.fill = STAT_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # ── Idle ports section (row 4+) ───────────────────────────────────────────
    # Blank row separator
    ws2.row_dimensions[3].height = 8

    idle_section_hdr = ws2.cell(row=4, column=1, value=f"Ports UP but Idle > 4 Weeks — {switch_name}")
    idle_section_hdr.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    idle_section_hdr.fill = PatternFill("solid", start_color="C00000")
    idle_section_hdr.border = THIN_BORDER
    ws2.merge_cells("A4:F4")
    ws2["A4"].alignment = CENTER

    idle_col_headers = ["Interface", "Description", "Connected", "Last Input", "Last Output", "MAC Address"]
    for col_idx, hdr in enumerate(idle_col_headers, start=1):
        cell = ws2.cell(row=5, column=col_idx, value=hdr)
        cell.font = SUBHDR_FONT
        cell.fill = SUBHDR_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    col_widths_2 = [20, 35, 12, 20, 20, 22]
    for i, w in enumerate(col_widths_2, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    idle_rows = [
        r for r in rows
        if r.get("connected") == "up"
        and is_idle_over_4_weeks(r.get("last_input", ""), r.get("last_output", ""))
    ]

    if idle_rows:
        for row_offset, r in enumerate(idle_rows, start=6):
            vals = [
                r.get("interface", ""),
                r.get("description", ""),
                r.get("connected", ""),
                r.get("last_input", ""),
                r.get("last_output", ""),
                r.get("end_device_mac", ""),
            ]
            for col_idx, val in enumerate(vals, start=1):
                cell = ws2.cell(row=row_offset, column=col_idx, value=val)
                cell.font = Font(name="Arial", size=10)
                cell.fill = IDLE_FILL
                cell.border = THIN_BORDER
    else:
        no_idle = ws2.cell(row=6, column=1, value="No ports found idle > 4 weeks.")
        no_idle.font = Font(italic=True, color="595959", name="Arial", size=10)
        no_idle.border = THIN_BORDER
        ws2.merge_cells("A6:F6")

    ws2.freeze_panes = "A6"

    wb.save(fname)

# -------------------------------- Core builder --------------------------------
def build_for_switch(sw: Dict[str, Any], cores: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    conn = connect(sw)
    try:
        if sw.get("secret"):
            try:
                conn.enable()
            except Exception:
                pass

        snmp_loc  = get_snmp_location(conn)

        swst      = conn.send_command("show interfaces status")
        sp        = conn.send_command("show interfaces switchport")
        intfs     = conn.send_command("show interfaces")
        descs     = conn.send_command("show interfaces description")
        poe       = conn.send_command("show power inline")
        mac_tab   = conn.send_command("show mac address-table")
        arp_local = conn.send_command("show ip arp")
        try:
            dhcp_snoop = conn.send_command("show ip dhcp snooping binding")
        except Exception:
            dhcp_snoop = ""

        base   = parse_interfaces_status(swst)
        spd    = parse_switchport_detail(sp)
        ifsum  = parse_interfaces_last_input_bulk(intfs)
        dmap   = parse_descriptions(descs)
        power  = parse_poe(poe)
        macs   = parse_mac_table(mac_tab)
        arpmap = parse_arp(arp_local)
        dstat  = parse_dhcp_snoop(dhcp_snoop)

        core_arp: Dict[str, str] = {}
        for core in cores:
            try:
                cconn = connect(core)
                try:
                    if core.get("secret"):
                        cconn.enable()
                except Exception:
                    pass
                cartxt = cconn.send_command("show ip arp")
                core_arp.update(parse_arp(cartxt))
            except Exception:
                pass
            finally:
                try:
                    cconn.disconnect()
                except Exception:
                    pass

        for intf, b in base.items():
            spinfo     = spd.get(intf, {})
            trunk_mode = spinfo.get("trunk_mode", False)
            vlans      = spinfo.get("vlans") or b.get("vlan_col", "")
            voice_vlan = spinfo.get("voice_vlan", "")
            desc       = dmap.get(intf, "")
            pwr        = power.get(intf, 0.0)

            li_str, lo_str = get_last_io_per_interface(conn, intf)
            if not li_str:
                li_str = ifsum.get(intf, {}).get("last_input", "")
            last_input  = li_str
            last_output = lo_str

            mac_list = macs.get(intf, [])
            if not mac_list:
                try:
                    per = conn.send_command(f"show mac address-table interface {intf}")
                    per_parsed = parse_mac_table(per)
                    mac_list = per_parsed.get(intf, [])
                except Exception:
                    pass

            ips: List[str] = []
            srcs: List[str] = []
            for mac in mac_list:
                ip = arpmap.get(mac) or core_arp.get(mac) or ""
                ips.append(ip)
                if mac in dstat:
                    srcs.append("DHCP")
                elif ip:
                    srcs.append("Static")
                else:
                    srcs.append("")

            max_sec_devices = ""
            try:
                ps = conn.send_command(f"show port-security interface {intf}")
                m = re.search(r"(?:Max(?:imum)?(?:\s+Secure)?\s+(?:MAC|Mac)\s+(?:Addr(?:esses)?|Addresses)|Max Secure Addrs)\s*:\s*(\d+)", ps, re.I)
                if m: max_sec_devices = m.group(1)
                if not max_sec_devices:
                    run = conn.send_command(f"show running-config interface {intf}")
                    m2 = re.search(r"switchport\s+port-security\s+maximum\s+(\d+)", run, re.I)
                    if m2: max_sec_devices = m2.group(1)
            except Exception:
                pass

            rows.append({
                "switch_name":     sw["name"],
                "switch_ip":       sw["host"],
                "snmp_location":   snmp_loc,
                "interface":       intf,
                "description":     desc,
                "connected":       "up" if b["connected"] else "down",
                "trunk_mode":      "trunk" if trunk_mode else "access",
                "vlans":           vlans,
                "voice_vlan":      voice_vlan,
                "duplex":          b["duplex"],
                "speed":           b["speed"],
                "power_watts":     pwr,
                "max_sec_devices": max_sec_devices,
                "last_input":      last_input,
                "last_output":     last_output,
                "end_device_mac":  ";".join(format_mac(m) for m in mac_list),
                "end_device_ip":   ";".join(ips),
                "ip_source":       ";".join(srcs),
            })

    finally:
        try:
            conn.disconnect()
        except Exception:
            pass

    return rows

# -------------------------------- Main --------------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--inventory", required=True)
    args = ap.parse_args()

    with open(args.inventory, "r", encoding="utf-8") as f:
        inv = yaml.safe_load(f)

    cores: List[Dict[str, Any]] = []
    for h in inv.get("cores", []):
        cores.append({
            "name": h["name"],
            "host": h["host"],
            "username": h.get("username", inv["credentials"]["username"]),
            "password": h.get("password", inv["credentials"]["password"]),
            "secret": h.get("secret", inv["credentials"].get("secret", "")),
            "device_type": h.get("device_type", "cisco_ios"),
        })

    outdir = "reports"
    os.makedirs(outdir, exist_ok=True)

    for sw in inv.get("switches", []):
        dev = {
            "name": sw["name"],
            "host": sw["host"],
            "username": sw.get("username", inv["credentials"]["username"]),
            "password": sw.get("password", inv["credentials"]["password"]),
            "secret": sw.get("secret", inv["credentials"].get("secret", "")),
            "device_type": sw.get("device_type", "cisco_ios"),
        }

        ts = time.strftime("%Y%m%d-%H%M%S")
        fname = os.path.join(outdir, f"{sw['name']}-{ts}.xlsx")

        rows = build_for_switch(dev, cores)
        write_xlsx(fname, rows, sw["name"])

        print(f"Wrote {fname}")

if __name__ == "__main__":
    main()